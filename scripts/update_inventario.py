"""Update data/Inventario_Manuales.xlsx from the on-disk corpus.

Scans local `Manuales_*` directories and rebuilds the per-fabricante sheets
plus the Resumen. Driven by a single `FABRICANTES` registry — adding a new
fabricante is a one-dict-entry change (see "HOW TO ADD A FABRICANTE" below).

Usage:
    python scripts/update_inventario.py              # update in-place
    python scripts/update_inventario.py --dry-run    # summarise only
    python scripts/update_inventario.py --only Morley,Notifier

Sheet schema (all fabricantes share this; column meanings below):
    Producto | Tipo documento | Idioma | Subcarpeta | Archivo | Tamaño (KB)

- `Producto`: extracted from filename via per-fabricante regex; falls back
  to filename stem.
- `Tipo documento`: classified by filename keywords (install/usuario/guía/
  troubleshoot/etc.) or overridden per-subfolder.
- `Idioma`: either fixed per-subfolder (e.g. EN_unico → "EN") or inferred
  from filename tokens.
- `Subcarpeta`: the directory origin — key for traceability and gap-hunting.
- `Archivo`, `Tamaño (KB)`: filename + size on disk.

===========================================================================
HOW TO ADD A NEW FABRICANTE (checklist for future sessions)
===========================================================================
1. Download manuals into one or more `Manuales_<Name>*` folders.
2. Add a new entry to the `FABRICANTES` dict below:
      "<Name>": {
          "subfolders": {"<key>": ROOT / "Manuales_<Name>_<variant>", ...},
          "model_regex": re.compile(r"...tailored regex...", re.IGNORECASE),
          "resumen_estado": "short state string for the Resumen sheet",
          "idioma_por_subcarpeta": {"<key>": "ES" or None, ...},  # None → heuristic
          "tipo_override_por_subcarpeta": {"<key>": "Guía troubleshooting", ...},  # optional
          "metadata_sidecar": ROOT / "path/to/_metadata.json",  # optional (Excel-sourced)
      }
3. Run `python scripts/update_inventario.py`. The sheet is rebuilt, the
   Resumen row upserted, a backup of the previous .xlsx goes to .bak.xlsx.
4. Verify the Resumen row manually (Productos / Documentos counts make sense).
5. Commit the .xlsx and the script changes together.
"""
from __future__ import annotations

import argparse
import json
import logging
import re
import shutil
import sys
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

logging.basicConfig(format="%(asctime)s - %(levelname)s - %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

ROOT = Path(__file__).resolve().parent.parent
INVENTARIO_PATH = ROOT / "data" / "Inventario_Manuales.xlsx"


# ---------------------------------------------------------------------------
# Regex libraries (one per fabricante)
# ---------------------------------------------------------------------------
MORLEY_MODEL_RX = re.compile(
    r"\b(ZX[e2-5rS]?[e]?|DXc[- ]?(?:Connexion)?|RP1[rR]?|MI[- ]?FLX[- ]?\d+|"
    r"MI-?(?:\d+|[A-Z]+)|MIE-?MI-?\d+|MIA[- ]?\d+|ECO10\d+|"
    r"F5000|FAAST[- ]?(?:LT|FLEX|ML)?|NFS\w*|"
    r"UCIP|TG\b|TG-?IP\w*|VSN\w*|CCM\w*|"
    r"MIW-?INT)\b",
    re.IGNORECASE,
)

NOTIFIER_MODEL_RX = re.compile(
    r"\b(AFP[- ]?\d+\w*|ID\d+\w*|AM\d+\w*|PEARL|INSPIRE|"
    r"System\s*5000|S5000\w*|VESDA[- ]?E\w*|VESDA\w*|"
    r"SDX[- ]?\d+\w*|NFS[- ]?\d+\w*|NFS2[- ]?\d+\w*|"
    r"FAAST[- ]?\w*|MS[- ]?\d+\w*|XP[- ]?\d+\w*|"
    r"RP[- ]?\d+\w*|FSP[- ]?\d+\w*|B\d{3}\w*|HOCHIKI[- ]?\w*)\b",
    re.IGNORECASE,
)

KIDDE_MODEL_RX = re.compile(
    r"\b(2X-AT[- ]?F\d\w*|2X-A[FRT]\w*|2X-A\b|NC-P[FX]\d\w*|NC\b|"
    r"KFP-\w+|ZP[12]-\w+|1X-[FX]\d\w*|2010-\w+|FP\d+\w*|FC\d+\w*|FEP\d+\w*)\b",
    re.IGNORECASE,
)


# ---------------------------------------------------------------------------
# Fabricantes registry — add new ones here.
# ---------------------------------------------------------------------------
FABRICANTES: dict = {
    "Morley": {
        "subfolders": {
            "publico": ROOT / "Manuales_Morley",
            "privado": ROOT / "Manuales_Morley_Privado",
            "guias": ROOT / "Manuales_Morley_Guias",
        },
        "model_regex": MORLEY_MODEL_RX,
        "resumen_estado": "Público + Privado + Guías",
        "idioma_por_subcarpeta": {
            "publico": None, "privado": None, "guias": "ES",
        },
        "tipo_override_por_subcarpeta": {"guias": "Guía troubleshooting"},
        "metadata_sidecar": ROOT / "Manuales_Morley_Guias" / "_metadata.json",
    },
    "Notifier": {
        "subfolders": {
            "ES": ROOT / "Manuales_Notifier" / "ES",
            "EN_unico": ROOT / "Manuales_Notifier" / "EN_unico",
            "MIXED": ROOT / "Manuales_Notifier" / "MIXED",
            "ES_traducido": ROOT / "Manuales_Notifier" / "ES_traducido",
            "privado": ROOT / "Manuales_Notifier_Privado",
        },
        "model_regex": NOTIFIER_MODEL_RX,
        "resumen_estado": "ES + EN_único + MIXED + Privado",
        "idioma_por_subcarpeta": {
            "ES": "ES", "EN_unico": "EN", "MIXED": "MULTI",
            "ES_traducido": "ES→traducido", "privado": None,
        },
        "tipo_override_por_subcarpeta": {},
        "metadata_sidecar": None,
    },
    "Kidde": {
        "subfolders": {"portal": ROOT / "Manuales_Kidde"},
        "model_regex": KIDDE_MODEL_RX,
        "resumen_estado": "Portal firesecurityproducts (paneles Control · ES+EN)",
        "idioma_por_subcarpeta": {"portal": None},
        "tipo_override_por_subcarpeta": {},
        "metadata_sidecar": ROOT / "Manuales_Kidde" / "_metadata.json",
    },
    # ADD A NEW FABRICANTE HERE — follow the schema above. Example:
    # "Bosch": {
    #     "subfolders": {"publico": ROOT / "Manuales_Bosch"},
    #     "model_regex": re.compile(r"\b(FPA\d+\w*|Avenar\w*)\b", re.IGNORECASE),
    #     "resumen_estado": "Público",
    #     "idioma_por_subcarpeta": {"publico": None},
    #     "tipo_override_por_subcarpeta": {},
    #     "metadata_sidecar": None,
    # },
}


# ---------------------------------------------------------------------------
# Tipo-documento heuristics (shared across fabricantes)
# ---------------------------------------------------------------------------
TIPO_PATTERNS = [
    (re.compile(r"\b(installation|instalaci[oó]n|install[- _])", re.IGNORECASE), "Manual instalación"),
    (re.compile(r"\b(user|usuario|utilizador|utilisateur)\b", re.IGNORECASE), "Manual usuario"),
    (re.compile(r"\b(commissioning|puesta en marcha)\b", re.IGNORECASE), "Puesta en marcha"),
    (re.compile(r"\b(quick|r[áa]pid[ao]|QIG|qig)\b", re.IGNORECASE), "Guía rápida"),
    (re.compile(r"\b(troubleshoot|averia|aver[íi]a|fault|diagn[oó]stico)\b", re.IGNORECASE), "Troubleshooting"),
    (re.compile(r"\b(programming|programaci[oó]n)\b", re.IGNORECASE), "Manual programación"),
    (re.compile(r"\b(product|producto|manual)\b", re.IGNORECASE), "Manual técnico"),
    (re.compile(r"\b(datasheet|data sheet|hoja|ficha)\b", re.IGNORECASE), "Datasheet"),
    (re.compile(r"\b(cert|certificat)\b", re.IGNORECASE), "Certificado"),
    (re.compile(r"\b(cat[áa]log|catalog)\b", re.IGNORECASE), "Catálogo"),
    (re.compile(r"\b(guia|gu[íi]a|guide|handbook)\b", re.IGNORECASE), "Guía"),
]


def classify_tipo(filename: str) -> str:
    for rx, label in TIPO_PATTERNS:
        if rx.search(filename):
            return label
    return "Otro"


def detect_language(filename: str) -> str:
    """Best-effort language detection from filename tokens."""
    lower = filename.lower()
    if re.search(r"\b(_fr|\bfrance|francais|français)\b|_fr\.", lower):
        return "FR"
    if re.search(r"\b(_pt|portug|portuguese)\b|_pt\.", lower):
        return "PT"
    if re.search(r"\b(_it|_ita|italiano|italian)\b|_ita?\.", lower):
        return "IT"
    if re.search(r"\bmultiling\b", lower):
        return "MULTI"
    if re.search(r"(_sp|_es|spanish|español|espa\xf1ol)\b", lower):
        return "ES"
    if re.search(r"(_en|english)\b", lower):
        return "EN"
    return "ES"  # default fallback for Spanish-distributor corpora


def load_metadata_sidecar(path: Path | None) -> dict[str, dict]:
    """Return filename → row-metadata mapping (empty if path missing)."""
    if not path or not path.exists():
        return {}
    try:
        entries = json.loads(path.read_text(encoding="utf-8"))
        return {e["local_filename"]: e for e in entries if isinstance(e, dict)}
    except Exception as exc:  # noqa: BLE001
        logger.warning("Failed to load metadata sidecar %s: %s", path, exc)
        return {}


def extract_product(filename: str, model_regex: re.Pattern,
                    metadata_equipo: str | None = None) -> str:
    """Pull a product model out of the filename. Prefer explicit metadata
    (e.g. Excel-sourced EQUIPO for troubleshooting guides) if provided."""
    if metadata_equipo:
        return metadata_equipo
    match = model_regex.search(filename)
    if match:
        return match.group(0).upper()
    return Path(filename).stem[:50]


def scan_fabricante(name: str, config: dict) -> list[dict]:
    """Scan all subfolders of one fabricante and return a list of row dicts."""
    rows = []
    sidecar = load_metadata_sidecar(config.get("metadata_sidecar"))
    model_regex: re.Pattern = config["model_regex"]
    idioma_map = config.get("idioma_por_subcarpeta", {})
    tipo_override = config.get("tipo_override_por_subcarpeta", {})

    for subcarpeta, folder in config["subfolders"].items():
        if not folder.exists():
            logger.info("  skip %s (folder missing)", folder.name)
            continue
        pdfs = sorted(folder.glob("*.pdf"))
        logger.info("  %s: %d PDFs", folder.name, len(pdfs))
        for pdf in pdfs:
            size_kb = max(1, round(pdf.stat().st_size / 1024))
            meta = sidecar.get(pdf.name, {})
            producto = extract_product(pdf.name, model_regex, meta.get("equipo"))

            # Tipo: sidecar (PIM-sourced) > per-subfolder override > filename heuristic.
            if meta.get("tipo"):
                tipo = meta["tipo"]
            elif subcarpeta in tipo_override:
                tipo = tipo_override[subcarpeta]
            else:
                tipo = classify_tipo(pdf.name)

            # Idioma: sidecar > per-subfolder fixed > filename heuristic.
            forced_idioma = idioma_map.get(subcarpeta)
            idioma = meta.get("idioma") or forced_idioma or detect_language(pdf.name)

            rows.append({
                "producto": producto,
                "tipo_documento": tipo,
                "idioma": idioma,
                "subcarpeta": subcarpeta,
                "archivo": pdf.name,
                "tamaño_kb": size_kb,
            })
    return rows


# ---------------------------------------------------------------------------
# Excel I/O
# ---------------------------------------------------------------------------

SHEET_HEADER = ["Producto", "Tipo documento", "Idioma", "Subcarpeta", "Archivo", "Tamaño (KB)"]
SHEET_COL_WIDTHS = [30, 22, 14, 14, 60, 12]


def rebuild_sheet(wb: openpyxl.Workbook, sheet_name: str, rows: list[dict]) -> None:
    """Create (or replace) a fabricante sheet with the standard 6-column schema."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(SHEET_HEADER)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
    for r in rows:
        ws.append([
            r["producto"], r["tipo_documento"], r["idioma"],
            r["subcarpeta"], r["archivo"], r["tamaño_kb"],
        ])
    for i, w in enumerate(SHEET_COL_WIDTHS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


def upsert_resumen_row(wb: openpyxl.Workbook, fabricante: str,
                       n_productos: int, n_docs: int, estado: str) -> None:
    """Upsert (fabricante, productos, documentos, estado) into the Resumen sheet."""
    if "Resumen" not in wb.sheetnames:
        logger.warning("Resumen sheet missing; cannot update summary")
        return
    ws = wb["Resumen"]

    # Find header row by looking for "Fabricante" in the first 10 rows.
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        if any(isinstance(c.value, str) and c.value.strip() == "Fabricante" for c in row):
            header_row = row[0].row
            break
    if header_row is None:
        logger.warning("Could not find 'Fabricante' header in Resumen; skipping")
        return

    # Find existing row for this fabricante; else first empty row below header.
    target_row = None
    first_empty = None
    for row in ws.iter_rows(min_row=header_row + 1, max_row=header_row + 50):
        val = row[0].value
        if val is None and first_empty is None:
            first_empty = row[0].row
        if isinstance(val, str) and val.strip().lower() == fabricante.lower():
            target_row = row[0].row
            break

    target_row = target_row or first_empty or (header_row + 1)
    ws.cell(row=target_row, column=1, value=fabricante)
    ws.cell(row=target_row, column=2, value=n_productos)
    ws.cell(row=target_row, column=3, value=n_docs)
    ws.cell(row=target_row, column=4, value=estado)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true",
                    help="Only summarise what would be written; don't touch the xlsx")
    ap.add_argument("--only", default=None,
                    help="Comma-separated list of fabricantes to process "
                         "(default: all in the FABRICANTES registry)")
    args = ap.parse_args()

    if args.only:
        selected = [name.strip() for name in args.only.split(",") if name.strip()]
        unknown = [n for n in selected if n not in FABRICANTES]
        if unknown:
            logger.error("Unknown fabricantes: %s (available: %s)",
                         unknown, list(FABRICANTES.keys()))
            return 2
    else:
        selected = list(FABRICANTES.keys())

    scanned = {}  # name → (rows, n_productos)
    for name in selected:
        logger.info("=== Scanning %s ===", name)
        rows = scan_fabricante(name, FABRICANTES[name])
        productos = {r["producto"] for r in rows}
        logger.info("%s: %d rows, %d productos, por subcarpeta=%s",
                    name, len(rows), len(productos),
                    dict(Counter(r["subcarpeta"] for r in rows)))
        scanned[name] = (rows, len(productos))

    if args.dry_run:
        logger.info("Dry-run, not writing.")
        return 0

    if not INVENTARIO_PATH.exists():
        logger.error("Inventario not found at %s", INVENTARIO_PATH)
        return 1

    wb = openpyxl.load_workbook(str(INVENTARIO_PATH))
    logger.info("Opened %s — existing sheets: %s", INVENTARIO_PATH.name, wb.sheetnames)

    for name, (rows, n_prod) in scanned.items():
        rebuild_sheet(wb, name, rows)
        upsert_resumen_row(
            wb, name, n_prod, len(rows),
            FABRICANTES[name]["resumen_estado"],
        )

    # Backup once (don't overwrite an existing backup — it's our undo).
    backup = INVENTARIO_PATH.with_suffix(".bak.xlsx")
    if not backup.exists():
        shutil.copy2(INVENTARIO_PATH, backup)
        logger.info("Backup written: %s", backup.name)

    wb.save(str(INVENTARIO_PATH))
    logger.info("Saved. Sheets now: %s", wb.sheetnames)
    return 0


if __name__ == "__main__":
    sys.exit(main())
